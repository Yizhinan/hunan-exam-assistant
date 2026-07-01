"""Excel parser for position import — multi-sheet, header auto-detection, field mapping."""

import re
import logging
from io import BytesIO
from typing import Any

import openpyxl

from sqlalchemy import select

from app.models.position import PositionHistory
from app.models.user import gen_uuid

logger = logging.getLogger(__name__)

# ============================================================
# Column header → internal key mapping (23 columns)
# ============================================================
HEADER_MAP: dict[str, str] = {
    "考区": "city_raw",
    "单位名称": "department",
    "用人单位名称": "department",          # 2021 格式
    "单位层级": "unit_level",
    "职位名称": "position_name",
    "职位性质": "position_nature",
    "笔试考试科目": "exam_subject",
    "招录人数": "enrollment_count",
    "报考人员身份要求": "identity_requirement",
    "基层工作年限要求": "experience_requirement",
    "性别要求": "gender_requirement",
    "最高年龄要求": "age_limit",
    "最低学历要求": "education_requirement",
    "学位要求": "degree_requirement",
    "专业要求": "major_requirement",
    "户籍要求": "household_requirement",
    "职业资格证书等要求": "cert_requirement",
    "其他招录条件及说明": "other_requirements",
    "最低服务年限要求": "service_year",
    "录用后在本单位最低服务年限": "service_year",  # 2021 格式
    "笔试最低开考比例": "exam_min_ratio",
    "体能测评项目和标准": "physical_test",
    "是否需体能测评": "physical_test",       # 2024 格式：布尔值
    "体检标准": "medical_standard",
    "咨询电话1": "phone1",
    "咨询电话2": "phone2",
}

# Columns required to consider a row valid
REQUIRED_COLUMNS = ["单位名称", "职位名称"]

# ============================================================
# Header detection
# ============================================================
def find_header_row(ws, max_scan: int = 5) -> int | None:
    """Scan first max_scan rows for the header row.

    Tries multiple header indicators: 考区 (2026 format), 单位名称, 序号 (2024 format).
    """
    for row_idx in range(1, max_scan + 1):
        for cell in ws[row_idx]:
            if cell.value:
                val = str(cell.value).strip()
                if val in ("考区", "单位名称", "序号"):
                    return row_idx
    return None

# ============================================================
# City normalization
# ============================================================
CITY_NORMALIZE: dict[str, str] = {
    "长沙市": "长沙", "株洲市": "株洲", "湘潭市": "湘潭", "衡阳市": "衡阳",
    "邵阳市": "邵阳", "岳阳市": "岳阳", "常德市": "常德", "张家界市": "张家界",
    "益阳市": "益阳", "郴州市": "郴州", "永州市": "永州", "怀化市": "怀化",
    "娄底市": "娄底", "湘西州": "湘西", "湘西自治州": "湘西",
    "省直": "省直",
}


# ============================================================
# City from sheet category (for files without 考区 column)
# ============================================================
_SHEET_NAME_TO_CITY: dict[str, str] = {
    "省直机关": "省直",
    "省直单位": "省直",                   # 2022 格式
    "省直机关戒毒单位": "省直",
    "省直监狱戒毒单位": "省直",           # 2021 格式
    # 带"市"后缀（2024 格式）
    "长沙市": "长沙", "株洲市": "株洲", "湘潭市": "湘潭",
    "衡阳市": "衡阳", "邵阳市": "邵阳", "岳阳市": "岳阳",
    "常德市": "常德", "张家界市": "张家界", "益阳市": "益阳",
    "郴州市": "郴州", "永州市": "永州", "怀化市": "怀化",
    "娄底市": "娄底", "湘西自治州": "湘西",
    # 不带"市"后缀（2021 格式，sheet 名即标准城市名）
    "长沙": "长沙", "株洲": "株洲", "湘潭": "湘潭",
    "衡阳": "衡阳", "邵阳": "邵阳", "岳阳": "岳阳",
    "常德": "常德", "张家界": "张家界", "益阳": "益阳",
    "郴州": "郴州", "永州": "永州", "怀化": "怀化",
    "娄底": "娄底", "湘西州": "湘西",
}


def _city_from_sheet_name(sheet_name: str) -> str | None:
    """Derive city from sheet name when 考区 column is absent (2024 format)."""
    s = sheet_name.strip()
    if s in _SHEET_NAME_TO_CITY:
        return _SHEET_NAME_TO_CITY[s]
    # Try partial match for city names in sheet name
    for sheet_city, norm_city in _SHEET_NAME_TO_CITY.items():
        if sheet_city in s:
            return norm_city
    return None


def normalize_city(raw: str) -> str:
    """Strip 市 suffix, handle special cases. Preserve 州 in city names like 永州/郴州."""
    raw = raw.strip()
    if raw in CITY_NORMALIZE:
        return CITY_NORMALIZE[raw]
    # Strip trailing 市 only (not 州 — 永州/郴州 have 州 as part of name)
    if raw.endswith("市") and len(raw) > 2:
        return raw[:-1]
    return raw


# ============================================================
# Category detection
# ============================================================
def detect_exam_category(exam_subject: str | None, unit_level: str | None) -> str:
    """
    Determine exam_category from exam_subject text + unit_level fallback.
    Returns: 行政执法 / 县乡基层 / 省市直
    """
    if exam_subject:
        s = exam_subject.strip()
        if "行政执法" in s:
            return "行政执法"
        if "县乡" in s or "乡镇" in s:
            return "县乡基层"
        if "省市" in s:
            return "省市直"

    # Fallback from unit_level
    if unit_level:
        u = unit_level.strip()
        if "县" in u or "乡镇" in u:
            return "县乡基层"
        if "市" in u or "省" in u:
            return "省市直"

    return "省市直"  # safe default


# ============================================================
# Education / degree normalization
# ============================================================
def normalize_education(raw: str | None) -> str:
    if not raw:
        return "本科及以上"
    r = raw.strip()
    if "研究生" in r or "硕士" in r:
        return "硕士研究生及以上"
    if "大专" in r or "专科" in r:
        return "大专及以上"
    return "本科及以上"


def normalize_degree(raw: str | None) -> str:
    if not raw:
        return "学士"
    r = raw.strip()
    if "博士" in r:
        return "博士"
    if "硕士" in r:
        return "硕士"
    if "学士" in r:
        return "学士"
    if "无" in r:
        return "无要求"
    return "学士"


def build_column_index(ws, header_row: int) -> dict[str, int]:
    """Map {internal_key: column_index} from header row."""
    col_index: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        val = str(cell.value).strip() if cell.value else ""
        if val in HEADER_MAP:
            col_index[HEADER_MAP[val]] = col_idx
    return col_index


# ============================================================
# Sheet name → org_category normalization
# ============================================================
def normalize_org_category(sheet_name: str) -> str:
    """Normalize Excel sheet name to org_category value.

    Handles two formats:
      - 2026: sheet names are org categories (省直机关及直属单位, 法院系统, ...)
      - 2024: sheet names are city names (长沙市, 省直机关, ...) → map to org categories
    """
    s = sheet_name.strip()
    s = s.replace("　", "").replace("\xa0", "")  # full-width spaces

    # 2026 format: direct org category names
    KNOWN = [
        "省直机关及直属单位",
        "市州及以下机关",
        "法院系统",
        "检察院系统",
        "公安系统",
        "综合行政执法队伍",
    ]
    for k in KNOWN:
        if k in s:
            return k

    # 2024/2022 format: city-based sheet names → map to org categories
    # Provincial-level sheets → 省直机关及直属单位
    if "省直" in s or "戒毒" in s:
        return "省直机关及直属单位"
    # City sheets → 市州及以下机关
    CITY_NAMES = [
        "长沙", "株洲", "湘潭", "衡阳", "邵阳", "岳阳", "常德",
        "张家界", "益阳", "郴州", "永州", "怀化", "娄底", "湘西",
    ]
    for c in CITY_NAMES:
        if c in s:
            return "市州及以下机关"

    return s  # fallback


# ============================================================
# Row parsing
# ============================================================
def parse_position_row_from_tuple(
    row_values: tuple, col_index: dict[str, int], year: int,
    org_category: str | None = None, sheet_name: str | None = None,
) -> dict[str, Any] | None:
    """Parse one row (as tuple of values) into a cleaned position dict."""

    def _get(key: str) -> str | None:
        ci = col_index.get(key)
        if ci is None or ci > len(row_values):
            return None
        # col_index uses 1-based indices, tuple is 0-based
        val = row_values[ci - 1]
        return str(val).strip() if val is not None else None

    city_raw = _get("city_raw")
    department = _get("department")
    position_name = _get("position_name")

    # If no 考区 column (2024 format), derive city from sheet name
    if not city_raw and sheet_name:
        city_raw = _city_from_sheet_name(sheet_name)

    if not city_raw or not department or not position_name:
        return None

    exam_subject = _get("exam_subject")
    unit_level = _get("unit_level")

    enrollment_raw = _get("enrollment_count")
    try:
        enrollment_count = int(float(enrollment_raw)) if enrollment_raw else 1
    except (ValueError, TypeError):
        enrollment_count = 1

    return {
        "year": year,
        "province": "湖南",
        "city": normalize_city(city_raw),
        "department": department,
        "position_name": position_name,
        "exam_category": detect_exam_category(exam_subject, unit_level),
        "org_category": org_category,
        "education_requirement": normalize_education(_get("education_requirement")),
        "degree_requirement": normalize_degree(_get("degree_requirement")),
        "major_requirement": _get("major_requirement") or "不限",
        "political_requirement": "不限",
        "gender_requirement": _get("gender_requirement") or "不限",
        "experience_requirement": _get("experience_requirement") or "不限",
        "age_limit": _get("age_limit") or "35周岁以下",
        "exam_subject": exam_subject,
        "enrollment_count": enrollment_count,
        "source": "用户导入",
        "is_active": True,
    }


# ============================================================
# Main parse function — processes all sheets
# ============================================================
def parse_position_excel(file_bytes: bytes, year: int) -> list[dict[str, Any]]:
    """
    Parse position Excel file (multi-sheet).
    Uses iter_rows for fast read_only iteration (avoids per-cell XML parsing).
    Returns list of cleaned position dicts ready for DB import.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    all_rows: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws, max_scan=5)

        if header_row is None:
            logger.warning("Sheet '%s': no header row found, skipping", sheet_name)
            continue

        col_index = build_column_index(ws, header_row)
        if "department" not in col_index:
            logger.warning("Sheet '%s': missing required column 单位名称, skipping", sheet_name)
            continue

        org_cat = normalize_org_category(sheet_name)
        sheet_count = 0
        # iter_rows with values_only=True is 50-100x faster than ws.cell()
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_values = tuple(row)
            try:
                parsed = parse_position_row_from_tuple(row_values, col_index, year, org_cat, sheet_name)
                if parsed:
                    all_rows.append(parsed)
                    sheet_count += 1
            except Exception:
                pass  # Skip malformed rows silently

        logger.info("Sheet '%s' → org_category='%s': parsed %d positions", sheet_name, org_cat, sheet_count)

    wb.close()
    return all_rows


# ============================================================
# DB import — handles both sync and async sessions
# ============================================================
async def import_positions_to_db(
    db, rows: list[dict[str, Any]], year: int
) -> dict:
    """
    Import parsed position rows into position_history table.
    Deduplicates by (department, position_name, year).
    Works with both sync (SQLite) and async (PostgreSQL) sessions.
    """
    created = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    is_async = not hasattr(db, "query")

    for i, row in enumerate(rows):
        dept = row["department"]
        pos_name = row["position_name"]

        try:
            # Check for existing record
            if is_async:
                result = await db.execute(
                    select(PositionHistory).where(
                        PositionHistory.department == dept,
                        PositionHistory.position_name == pos_name,
                        PositionHistory.year == year,
                    )
                )
                existing = result.scalar_one_or_none()
            else:
                existing = (
                    db.query(PositionHistory)
                    .filter(
                        PositionHistory.department == dept,
                        PositionHistory.position_name == pos_name,
                        PositionHistory.year == year,
                    )
                    .first()
                )

            if existing:
                # Update all fields except id
                for key, value in row.items():
                    if hasattr(existing, key):
                        setattr(existing, key, value)
                updated += 1
            else:
                new_pos = PositionHistory(id=gen_uuid(), **row)
                db.add(new_pos)
                created += 1

            # Commit every 100 rows
            if (i + 1) % 100 == 0:
                if is_async:
                    await db.commit()
                else:
                    db.commit()

        except Exception as e:
            skipped += 1
            errors.append({
                "row": i + 1,
                "message": str(e),
                "data": {"department": dept, "position_name": pos_name},
            })

    # Final commit
    if is_async:
        await db.commit()
    else:
        db.commit()

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors[:50],  # cap at 50
    }
