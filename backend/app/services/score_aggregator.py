"""Interview score Excel parser — groups scores by position, aggregates min/max/avg."""

import logging
from collections import defaultdict
from io import BytesIO
from statistics import mean
from typing import Any

import openpyxl

from sqlalchemy import select

from app.models.position import PositionHistory

logger = logging.getLogger(__name__)

# ============================================================
# Column header mapping for interview score Excel
# ============================================================
SCORE_HEADER_MAP: dict[str, str] = {
    "序号": "seq",
    "姓名": "name",
    "准考证号": "exam_id",
    "单位名称": "department",
    "职位名称": "position_name",
    "行政职业能力测验成绩": "xingce_score",
    "行政职业能力测验": "xingce_score",       # 简写版本
    "申论成绩": "shenlun_score",
    "申论": "shenlun_score",                  # 简写版本
    "专业成绩": "professional_score",
    "专业": "professional_score",               # 2022 简写版本
    "法律成绩": "law_score",                   # 法律专业科目
    "招警成绩": "police_score",               # 招警专业科目
    "财经成绩": "finance_score",              # 财经专业科目
    "笔试综合成绩": "written_total_score",
    "笔试总成绩": "written_total_score",       # 简写版本
    "笔试排名": "written_rank",
}

# Columns that contribute to professional/total score (summed together)
_PROFESSIONAL_SCORE_KEYS = ("professional_score", "law_score", "police_score", "finance_score")


def find_score_header_row(ws, max_scan: int = 5) -> int | None:
    """Scan first rows for the header containing 序号 or 姓名."""
    for row_idx in range(1, max_scan + 1):
        for cell in ws[row_idx]:
            if cell.value and "序号" in str(cell.value).strip():
                return row_idx
    # Fallback: try 姓名
    for row_idx in range(1, max_scan + 1):
        for cell in ws[row_idx]:
            if cell.value and "姓名" in str(cell.value).strip():
                return row_idx
    return None


def build_score_column_index(ws, header_row: int) -> dict[str, int]:
    """Map {internal_key: column_index} from score header row."""
    col_index: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        val = str(cell.value).strip() if cell.value else ""
        if val in SCORE_HEADER_MAP:
            col_index[SCORE_HEADER_MAP[val]] = col_idx
    return col_index


def parse_score_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Parse interview score Excel (single sheet).
    Returns list of individual score rows with department + position_name + total_score.
    Uses iter_rows(values_only=True) for performance.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    all_rows: list[dict[str, Any]] = []

    # Use first sheet
    ws = wb[wb.sheetnames[0]]
    header_row = find_score_header_row(ws, max_scan=5)

    if header_row is None:
        wb.close()
        raise ValueError("未找到表头行（需包含「序号」或「姓名」列）")

    col_index = build_score_column_index(ws, header_row)
    if "department" not in col_index or "position_name" not in col_index:
        wb.close()
        raise ValueError("缺少必要列：单位名称、职位名称")

    has_score_col = "written_total_score" in col_index
    # Pre-compute column indices for fast tuple-based access (col_index is 1-based)
    dept_ci = col_index["department"] - 1
    pos_ci = col_index["position_name"] - 1
    score_ci = col_index["written_total_score"] - 1 if has_score_col else -1
    xc_ci = col_index["xingce_score"] - 1 if "xingce_score" in col_index else -1
    sc_ci = col_index["shenlun_score"] - 1 if "shenlun_score" in col_index else -1
    # Professional score columns (0-based indices)
    prof_cis = [col_index[k] - 1 for k in _PROFESSIONAL_SCORE_KEYS if k in col_index]

    # Use iter_rows for O(n) performance (vs O(n²) ws.cell())
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        dept_val = row[dept_ci] if dept_ci < len(row) else None
        pos_val = row[pos_ci] if pos_ci < len(row) else None

        if not dept_val or not pos_val:
            continue

        department = str(dept_val).strip()
        position_name = str(pos_val).strip()

        # Parse total score
        total_score = None
        if has_score_col and score_ci < len(row):
            score_val = row[score_ci]
            if score_val is not None:
                try:
                    total_score = float(score_val)
                except (ValueError, TypeError):
                    continue

        if total_score is None:
            # Try to compute from components
            xc = sc = pc = 0.0
            if xc_ci >= 0 and xc_ci < len(row):
                try:
                    xc = float(row[xc_ci]) if row[xc_ci] else 0
                except (ValueError, TypeError):
                    xc = 0
            if sc_ci >= 0 and sc_ci < len(row):
                try:
                    sc = float(row[sc_ci]) if row[sc_ci] else 0
                except (ValueError, TypeError):
                    sc = 0
            for pci in prof_cis:
                if pci < len(row):
                    try:
                        pc += float(row[pci]) if row[pci] else 0
                    except (ValueError, TypeError):
                        pass
            total_score = xc + sc + pc
            if total_score == 0:
                continue

        all_rows.append({
            "department": department,
            "position_name": position_name,
            "written_total_score": total_score,
        })

    wb.close()
    logger.info("Parsed %d score entries from Excel", len(all_rows))
    return all_rows


def aggregate_scores(rows: list[dict]) -> list[dict[str, Any]]:
    """
    Group score rows by (department, position_name) and compute aggregates.
    Returns list of {department, position_name, min_score, max_score, avg_score, applicant_count}.
    """
    groups: dict[tuple[str, str], list[float]] = defaultdict(list)

    for row in rows:
        key = (row["department"], row["position_name"])
        score = row.get("written_total_score")
        if score is not None:
            groups[key].append(float(score))

    results = []
    for (dept, pos_name), scores in groups.items():
        results.append({
            "department": dept,
            "position_name": pos_name,
            "min_score_interview": round(min(scores), 2),
            "max_score_interview": round(max(scores), 2),
            "avg_score_interview": round(mean(scores), 1),
            "applicant_count": len(scores),
        })

    logger.info("Aggregated %d position groups from %d score entries", len(results), len(rows))
    return results


async def import_scores_to_db(
    db, aggregated: list[dict[str, Any]], year: int
) -> dict:
    """
    Match aggregated scores to position_history and update score fields.
    Works with both sync (SQLite) and async (PostgreSQL) sessions.
    """
    updated = 0
    not_found = 0
    errors: list[dict] = []

    is_async = not hasattr(db, "query")

    for i, ag in enumerate(aggregated):
        dept = ag["department"]
        pos_name = ag["position_name"]

        try:
            # Find matching position
            if is_async:
                result = await db.execute(
                    select(PositionHistory).where(
                        PositionHistory.department == dept,
                        PositionHistory.position_name == pos_name,
                        PositionHistory.year == year,
                    )
                )
                positions = result.scalars().all()
            else:
                positions = (
                    db.query(PositionHistory)
                    .filter(
                        PositionHistory.department == dept,
                        PositionHistory.position_name == pos_name,
                        PositionHistory.year == year,
                    )
                    .all()
                )

            if positions:
                for pos in positions:
                    pos.min_score_interview = ag["min_score_interview"]
                    pos.max_score_interview = ag["max_score_interview"]
                    pos.avg_score_interview = ag["avg_score_interview"]
                    pos.applicant_count = ag["applicant_count"]
                updated += len(positions)
            else:
                not_found += 1

            # Commit every 50 groups
            if (i + 1) % 50 == 0:
                if is_async:
                    await db.commit()
                else:
                    db.commit()

        except Exception as e:
            errors.append({
                "department": dept,
                "position_name": pos_name,
                "message": str(e),
            })

    # Final commit
    if is_async:
        await db.commit()
    else:
        db.commit()

    return {
        "updated": updated,
        "not_found": not_found,
        "total_groups": len(aggregated),
        "errors": errors[:50],
    }
