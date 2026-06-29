"""
完整导入 2023-2026 湖南省考真实数据。

数据来源：红星网 (hxw.gov.cn)
  topic/227 = 2023年
  topic/247 = 2024年
  topic/267 = 2025年
  topic/294 = 2026年

每年度抓取：
  1. 职位表 Excel
  2. 面试名单 Excel（含笔试成绩）
"""

import os, sys, re, json, urllib.request, time
from pathlib import Path
from collections import defaultdict

os.chdir(Path(__file__).resolve().parent.parent)
sys.path.insert(0, ".")

from app.core.database import engine, Base, SessionLocal
from app.models.position import PositionHistory
import app.models.user, app.models.document, app.models.essay, app.models.daily_essay  # noqa

import openpyxl

DATA_DIR = Path("D:/tmp/hunan_exam_data")
DATA_DIR.mkdir(exist_ok=True)

TOPIC_IDS = {2023: 227, 2024: 247, 2025: 267, 2026: 294}

def fetch_page(topic_id: int) -> str:
    """Fetch a topic page and return all content/article links plus attachment URLs."""
    url = f"https://www.hxw.gov.cn/topic/{topic_id}/index.html"
    print(f"  Fetching {url}...")
    req = urllib.request.Request(url, headers={"User-Agent": "HunanExamAssistant/1.0"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        html = resp.read().decode("utf-8", errors="replace")
    return html


def find_attachments(html: str, year: int) -> dict:
    """
    From a topic page, find:
      - Position table Excel (职位表)
      - Interview list Excel (面试名单, 资格审查 with scores)
    Returns {"position": url, "interview": url or None}
    """
    # Find all Excel/doc links
    links = re.findall(r"""href=["']([^"']*(?:xlsx|xls|docx|doc)[^"']*)["']""", html, re.I)
    # Find all content page links
    content_links = re.findall(r"""href=["'](/content/\d{4}/\d{2}/\d{2}/\d+\.html)["']""", html)

    result = {"position": None, "interview_excels": [], "content_pages": []}

    for link in links:
        full = link if link.startswith("http") else f"https://img.hxw.gov.cn{link}" if link.startswith("/") else f"https://www.hxw.gov.cn/{link}"
        # Position tables are usually published in January
        if f"/{year}/01-" in full or f"/{year-1}/12-" in full:
            if "xlsx" in full.lower() and "职位" not in full.lower():
                result["position"] = full
                print(f"    Position Excel: {full.split('/')[-1][:50]}")

    # Find interview-related pages (资格审查 or 面试公告)
    for cl in set(content_links):
        full = f"https://www.hxw.gov.cn{cl}"
        result["content_pages"].append(full)

    # For each content page, check if it's interview-related
    for page_url in result["content_pages"][:20]:
        try:
            time.sleep(0.5)
            req = urllib.request.Request(page_url, headers={"User-Agent": "HunanExamAssistant/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                page_html = resp.read().decode("utf-8", errors="replace")

            title_match = re.search(r'<title>([^<]+)</title>', page_html)
            title = title_match.group(1) if title_match else ""

            # Check if this is an interview/qualification review page
            is_interview = any(kw in title for kw in ["面试", "资格复审", "资格审查", "体能测评", "分数线"])

            if is_interview:
                page_links = re.findall(r"""href=["']([^"']*(?:xlsx|xls)[^"']*)["']""", page_html, re.I)
                for pl in page_links:
                    full_pl = pl if pl.startswith("http") else f"https://img.hxw.gov.cn{pl}" if pl.startswith("/") else f"https://www.hxw.gov.cn/{pl}"
                    result["interview_excels"].append(full_pl)
                    print(f"    Interview Excel ({title[:40]}): {full_pl.split('/')[-1][:50]}")
        except Exception as e:
            pass

    return result


def parse_position_excel(filepath: Path, year: int) -> list[dict]:
    """Parse a position table Excel."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    positions = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        header_row = None
        for r in range(1, min(8, ws.max_row + 1)):
            c1 = str(ws.cell(r, 1).value or "").strip()
            c2 = str(ws.cell(r, 2).value or "").strip()
            if c1 in ("地区", "说明") or c2 == "单位名称":
                header_row = r
                break
        if header_row is None:
            continue

        col_map = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(header_row, c).value or "").strip()
            col_map[h] = c

        def g(row, col_idx=None, *names):
            if col_idx:
                val = ws.cell(row, col_idx).value
                return str(val).strip() if val else ""
            for n in names:
                if n in col_map:
                    val = ws.cell(row, col_map[n]).value
                    return str(val).strip() if val else ""
            return ""

        for r in range(header_row + 1, ws.max_row + 1):
            city = g(r, 1)
            unit = g(r, None, "单位名称")
            pn = g(r, None, "职位名称")
            if not unit or not pn:
                continue
            if city in ("地区", "说明", "") and unit == "单位名称":
                continue

            ul = g(r, None, "单位层级")
            es = g(r, None, "笔试考试科目")
            category = "行政执法" if "行政执法" in es else (
                "省市直" if ul in ("省直机关", "省直", "市州直机关", "市州直") else "县乡基层"
            )

            enr_str = g(r, None, "招录人数")
            try: enr = int(float(enr_str))
            except: enr = 1

            major = g(r, None, "专业要求")
            positions.append({
                "year": year, "city": city, "department": unit,
                "position_name": pn, "exam_category": category,
                "education_requirement": g(r, None, "最低学历要求") or "本科及以上",
                "degree_requirement": g(r, None, "学位要求") or "学士",
                "major_requirement": major if major and major != "不限" else None,
                "gender_requirement": g(r, None, "性别要求") or "不限",
                "age_limit": g(r, None, "最高年龄要求") or "35周岁以下",
                "experience_requirement": g(r, None, "基层工作年限要求") or "不限",
                "applicant_identity": g(r, None, "报考人员身份要求") or "不限",
                "hukou_requirement": g(r, None, "户籍要求") or "不限",
                "cert_requirement": g(r, None, "职业资格证书等要求") or "不限",
                "other_conditions": g(r, None, "其他招录条件及说明") or "",
                "enrollment_count": enr,
                "interview_ratio": g(r, None, "面试与考察比例") or "1:3",
                "exam_subject": es,
                "unit_level": ul,
                "position_type": g(r, None, "职位性质"),
                "physical_test": g(r, None, "体能测评项目和标准") or "否",
                "min_service_years": g(r, None, "最低服务年限要求") or "不限",
            })

    wb.close()
    return positions


def parse_interview_excel(filepath: Path) -> dict[tuple[str, str], dict]:
    """
    Parse an interview list Excel.
    Returns {(unit, position_name): {min, max, avg, count}}
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    scores = defaultdict(list)

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 2:
            continue
        # Detect columns
        unit_col = pos_col = score_col = None
        for c in range(1, min(20, ws.max_column + 1)):
            h = str(ws.cell(1, c).value or "")
            if "招录单位" in h or "单位" in h:
                unit_col = c
            elif "招录职位" in h or "职位" in h:
                pos_col = c
            elif "笔试成绩" in h or "成绩" in h:
                score_col = c

        if not unit_col or not pos_col:
            # Try common positions
            unit_col, pos_col = 5, 6
            for c in range(1, 12):
                h = str(ws.cell(1, c).value or "")
                if "单位" in h: unit_col = c
                if "职位" in h: pos_col = c
                if "成绩" in h and score_col is None: score_col = c
        if not score_col:
            score_col = 11  # Default

        for r in range(2, ws.max_row + 1):
            unit = str(ws.cell(r, unit_col).value or "").strip()
            pn = str(ws.cell(r, pos_col).value or "").strip()
            sc = ws.cell(r, score_col).value
            if unit and pn and sc:
                try: scores[(unit, pn)].append(float(sc))
                except: pass

    wb.close()

    return {
        k: {"count": len(v), "min": round(min(v), 2), "max": round(max(v), 2),
            "avg": round(sum(v) / len(v), 2)}
        for k, v in scores.items()
    }


# ---- Main ----
def main():
    print("=" * 60)
    print("湖南省考多年数据导入")
    print("=" * 60)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    for year in [2023, 2024, 2025, 2026]:
        print(f"\n{'='*40}")
        print(f"  {year}年")
        print(f"{'='*40}")

        # Check if already imported
        existing = db.query(PositionHistory).filter(PositionHistory.year == year).count()
        if existing > 50:
            print(f"  Already have {existing} positions, skipping")
            continue

        topic_id = TOPIC_IDS[year]
        try:
            html = fetch_page(topic_id)
            assets = find_attachments(html, year)

            # Download position table
            pos_file = DATA_DIR / f"positions_{year}.xlsx"
            if assets["position"] and not pos_file.exists():
                print(f"  Downloading position table...")
                urllib.request.urlretrieve(assets["position"], str(pos_file))

            # Parse positions
            if pos_file.exists():
                positions = parse_position_excel(pos_file, year)
                print(f"  Parsed {len(positions)} positions")
            else:
                print(f"  No position table found, skipping")
                continue

            # Try to find and parse interview score data
            score_stats = {}
            for ie_url in assets.get("interview_excels", [])[:5]:
                ie_file = DATA_DIR / f"interview_{year}_{abs(hash(ie_url))}.xlsx"
                if not ie_file.exists():
                    try:
                        urllib.request.urlretrieve(ie_url, str(ie_file))
                    except:
                        continue

                if ie_file.exists():
                    try:
                        stats = parse_interview_excel(ie_file)
                        score_stats.update(stats)
                        print(f"    +{len(stats)} positions with scores from {ie_file.name[:40]}")
                    except Exception as e:
                        print(f"    Parse error: {e}")

            # Merge and import
            imported = 0
            with_scores = 0
            for p in positions:
                key = (p["department"], p["position_name"])
                sc = score_stats.get(key, {})

                # Also try fuzzy matching
                if not sc:
                    for k, v in score_stats.items():
                        if p["department"] in k[0] or k[0] in p["department"]:
                            if p["position_name"] in k[1] or k[1] in p["position_name"]:
                                sc = v
                                break

                if sc:
                    with_scores += 1

                db.add(PositionHistory(
                    year=p["year"], city=p["city"], department=p["department"],
                    position_name=p["position_name"], exam_category=p["exam_category"],
                    education_requirement=p["education_requirement"],
                    degree_requirement=p["degree_requirement"],
                    major_requirement=p["major_requirement"],
                    gender_requirement=p["gender_requirement"],
                    age_limit=p["age_limit"],
                    experience_requirement=p["experience_requirement"],
                    political_requirement="不限",  # Not a separate column in position table
                    enrollment_count=p["enrollment_count"],
                    interview_ratio=p["interview_ratio"],
                    min_score_interview=sc.get("min"),
                    max_score_interview=sc.get("max"),
                    avg_score_interview=sc.get("avg"),
                    source=f"红星网 topic/{topic_id}",
                ))
                imported += 1

            db.commit()
            print(f"  Imported: {imported} positions ({with_scores} with scores)")

        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()

    total = db.query(PositionHistory).count()
    print(f"\n{'='*40}")
    print(f"Total positions in DB: {total}")
    years = db.execute(
        "SELECT year, count(*) FROM position_history GROUP BY year ORDER BY year"
    ).fetchall()
    for y, c in years:
        scored = db.query(PositionHistory).filter(
            PositionHistory.year == y, PositionHistory.min_score_interview.isnot(None)
        ).count()
        print(f"  {y}: {c} positions ({scored} with scores)")
    db.close()


if __name__ == "__main__":
    main()
