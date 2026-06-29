"""
从红星网导入 2026 湖南省考真实职位数据和进面分数。

数据来源：
  职位表: https://img.hxw.gov.cn/2026/01-13/c8bf0670-4ac0-4d1b-acc8-5c2c06849917.xlsx
  面试名单: https://img.hxw.gov.cn/2026/04-21/57e0f6f7-161d-4adc-92ab-0d6f343684fb.xlsx

用法: cd backend && python scripts/import_real_data.py
"""

import os, sys, urllib.request, re
from pathlib import Path

os.chdir(Path(__file__).resolve().parent.parent)
sys.path.insert(0, ".")

from app.core.database import engine, Base, SessionLocal
from app.models.position import PositionHistory, UserProfile

# Import all models to ensure tables exist
import app.models.user  # noqa
import app.models.document  # noqa
import app.models.essay  # noqa
import app.models.daily_essay  # noqa

import openpyxl

# ---- Download files ----
DATA_DIR = Path("/tmp/hunan_exam_data")
DATA_DIR.mkdir(exist_ok=True)

POSITION_URL = "https://img.hxw.gov.cn/2026/01-13/c8bf0670-4ac0-4d1b-acc8-5c2c06849917.xlsx"
INTERVIEW_URL = "https://img.hxw.gov.cn/2026/04-21/57e0f6f7-161d-4adc-92ab-0d6f343684fb.xlsx"

pos_file = DATA_DIR / "position_2026.xlsx"
int_file = DATA_DIR / "interview_2026.xlsx"

for url, path in [(POSITION_URL, pos_file), (INTERVIEW_URL, int_file)]:
    if not path.exists():
        print(f"Downloading {url.split('/')[-1]}...")
        urllib.request.urlretrieve(url, str(path))
        print(f"  OK: {path.name}")

# ---- Parse interview scores FIRST (need per-position stats) ----
print("\n=== Parsing interview scores ===")
wb = openpyxl.load_workbook(str(int_file), data_only=True)

# Aggregate per (unit, position): [scores]
from collections import defaultdict

position_scores: dict[tuple[str, str], list[float]] = defaultdict(list)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(2, ws.max_row + 1):
        unit = str(ws.cell(r, 5).value or "").strip()
        pos_name = str(ws.cell(r, 6).value or "").strip()
        total_score = ws.cell(r, 11).value  # 笔试成绩 (column 11)

        if unit and pos_name and total_score:
            try:
                score = float(total_score)
                key = (unit, pos_name)
                position_scores[key].append(score)
            except (ValueError, TypeError):
                pass

wb.close()

# Calculate stats
score_stats: dict[tuple[str, str], dict] = {}
for key, scores in position_scores.items():
    score_stats[key] = {
        "count": len(scores),
        "min_score": round(min(scores), 2),
        "max_score": round(max(scores), 2),
        "avg_score": round(sum(scores) / len(scores), 2),
    }

print(f"Parsed {len(score_stats)} unique positions with interview scores")
print(f"Total candidates: {sum(len(v) for v in position_scores.values())}")

# ---- Parse position table ----
print("\n=== Parsing position table ===")
wb = openpyxl.load_workbook(str(pos_file), data_only=True)

def infer_category(unit_level: str, exam_subject: str) -> str:
    """Infer exam category from EXAM SUBJECT (most reliable)."""
    if not exam_subject:
        return "省市直"
    if "行政执法" in exam_subject:
        return "行政执法"
    if "县乡" in exam_subject:
        return "县乡基层"
    if "省市" in exam_subject:
        return "省市直"
    return "省市直"


positions = []
total_rows = 0
skipped = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  Sheet: {sheet_name} ({ws.max_row} rows)")

    # Find header row (look for the row containing 地区/单位名称)
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        first_cell = str(ws.cell(r, 1).value or "")
        second_cell = str(ws.cell(r, 2).value or "")
        if first_cell in ("地区", "说明") or second_cell in ("单位名称",):
            header_row = r
            break

    if header_row is None:
        print(f"    Skipping sheet {sheet_name}: no header found")
        continue

    # Map columns by header name
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(header_row, c).value or "").strip()
        col_map[h] = c

    def get_col(row, *names):
        for name in names:
            if name in col_map:
                val = ws.cell(row, col_map[name]).value
                return str(val).strip() if val else ""
        return ""

    for r in range(header_row + 1, ws.max_row + 1):
        city = str(ws.cell(r, 1).value or "").strip()  # Column 1 = city always
        unit = get_col(r, "单位名称")
        unit_level = get_col(r, "单位层级")
        pos_name = get_col(r, "职位名称")
        pos_type = get_col(r, "职位性质")
        exam_subject = get_col(r, "笔试考试科目")
        enrollment = get_col(r, "招录人数")
        identity_req = get_col(r, "报考人员身份要求")
        work_exp = get_col(r, "基层工作年限要求")
        gender = get_col(r, "性别要求")
        max_age = get_col(r, "最高年龄要求")
        education = get_col(r, "最低学历要求")
        degree = get_col(r, "学位要求")
        major = get_col(r, "专业要求")
        political = get_col(r, "户籍要求")  # 户籍 in this table
        cert = get_col(r, "职业资格证书等要求")
        other = get_col(r, "其他招录条件及说明")
        service_years = get_col(r, "最低服务年限要求")
        interview_ratio = get_col(r, "面试与考察比例")
        physical_test = get_col(r, "体能测评项目和标准")
        medical = get_col(r, "体检标准")
        phone1 = get_col(r, "咨询电话1")

        if not unit or not pos_name:
            continue

        # Parse enrollment
        try:
            enr = int(float(enrollment))
        except (ValueError, TypeError):
            enr = 1

        # Category
        category = infer_category(unit_level, exam_subject)

        # Check if we have interview scores for this position
        key = (unit, pos_name)
        scores = score_stats.get(key, {})

        # Also try fuzzy matching (some names might differ slightly)
        if not scores:
            for k, v in score_stats.items():
                if unit in k[0] or k[0] in unit:
                    if pos_name in k[1] or k[1] in pos_name:
                        scores = v
                        key = k
                        break

        positions.append({
            "year": 2026,
            "city": city,
            "department": unit,
            "position_name": pos_name,
            "exam_category": category,
            "exam_subject": exam_subject,
            "education_requirement": education or "本科及以上",
            "degree_requirement": degree or "学士",
            "major_requirement": major if major and major != "不限" else None,
            "political_requirement": political if political and "党员" in political else "不限",
            "gender_requirement": gender if gender and gender != "不限" else "不限",
            "experience_requirement": work_exp if work_exp and work_exp != "不限" else "不限",
            "age_limit": max_age if max_age else "35周岁以下",
            "enrollment_count": enr,
            "interview_ratio": interview_ratio if interview_ratio else "1:3",
            "min_score_interview": scores.get("min_score"),
            "max_score_interview": scores.get("max_score"),
            "avg_score_interview": scores.get("avg_score"),
        })
        total_rows += 1

wb.close()

print(f"\nParsed {total_rows} positions")

# ---- Import to database ----
print("\n=== Importing to database ===")
Base.metadata.create_all(bind=engine)
db = SessionLocal()

# Delete old 2026 data
old_count = db.query(PositionHistory).filter(PositionHistory.year == 2026).count()
if old_count > 0:
    db.query(PositionHistory).filter(PositionHistory.year == 2026).delete()
    print(f"Deleted {old_count} old 2026 records")

imported = 0
with_scores = 0
for p in positions:
    if p["min_score_interview"]:
        with_scores += 1

    db.add(PositionHistory(
        year=p["year"], city=p["city"],
        department=p["department"], position_name=p["position_name"],
        exam_category=p["exam_category"],
        exam_subject=p.get("exam_subject"),
        education_requirement=p["education_requirement"],
        degree_requirement=p["degree_requirement"],
        major_requirement=p["major_requirement"],
        political_requirement=p["political_requirement"],
        gender_requirement=p["gender_requirement"],
        experience_requirement=p["experience_requirement"],
        age_limit=p["age_limit"],
        enrollment_count=p["enrollment_count"],
        interview_ratio=p["interview_ratio"],
        min_score_interview=p["min_score_interview"],
        max_score_interview=p["max_score_interview"],
        avg_score_interview=p["avg_score_interview"],
    ))
    imported += 1

db.commit()

# Stats
total = db.query(PositionHistory).count()
print(f"\nImported: {imported} positions ({with_scores} with real scores)")
print(f"Total DB positions: {total}")
db.close()

# ---- Show sample ----
print("\n=== Sample (岳阳监狱 related) ===")
for key, stats in score_stats.items():
    if "监狱" in key[0] or "监狱" in key[1]:
        print(f"  {key[0]} / {key[1]}: min={stats['min_score']} max={stats['max_score']} avg={stats['avg_score']} ({stats['count']}人)")
