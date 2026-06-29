"""
完整导入 2023-2026 湖南省考职位数据 V3
支持不同年份的Excel格式差异：
  - 2026: 列1=城市, 表头在R1-R2, 含"笔试考试科目"
  - 2023/2024: 每个sheet是一个城市(或系统), 列1=序号, 城市从sheet名提取
  - 2025: 类似2026格式
  - 法院/检察院系统: 独立文件, 不同表头布局

用法: cd backend && python scripts/import_all_years_v3.py
"""

import os, sys, re, urllib.request, time
from pathlib import Path
from collections import defaultdict

os.chdir(Path(__file__).resolve().parent.parent)
sys.path.insert(0, ".")

from app.core.database import engine, Base, SessionLocal
from app.models.position import PositionHistory
import app.models.user, app.models.document, app.models.essay, app.models.daily_essay  # noqa

import openpyxl

DATA_DIR = Path("D:/tmp/hunan_exam_data")
DATA_DIR.mkdir(exist_ok=True)

TOPIC_IDS = {2023: 227, 2024: 247, 2025: 267, 2026: 294}

# ============================================================
# Helpers
# ============================================================
def infer_category(exam_subject: str) -> str:
    if not exam_subject:
        return "省市直"
    if "行政执法" in exam_subject:
        return "行政执法"
    if "县乡" in exam_subject:
        return "县乡基层"
    if "省市" in exam_subject:
        return "省市直"
    return "省市直"


def city_from_sheet_name(sheet_name: str) -> str | None:
    """Extract city code from sheet name. E.g. '长沙市' -> '长沙', '省直单位' -> '省直'"""
    # Direct matches
    city_map = {
        "省直单位": "省直", "省直机关": "省直", "省直": "省直",
        "省直监狱戒毒单位": "省直", "省直机关及直属单位": "省直",
        "长沙市": "长沙", "株洲市": "株洲", "湘潭市": "湘潭",
        "衡阳市": "衡阳", "邵阳市": "邵阳", "岳阳市": "岳阳",
        "常德市": "常德", "张家界市": "张家界", "益阳市": "益阳",
        "郴州市": "郴州", "永州市": "永州", "怀化市": "怀化",
        "娄底市": "娄底", "湘西自治州": "湘西", "湘西州": "湘西",
        "法院系统": "省直", "检察院系统": "省直",
    }
    if sheet_name in city_map:
        return city_map[sheet_name]
    for k, v in city_map.items():
        if k in sheet_name or sheet_name in k:
            return v
    return None


def find_header_row(ws, required_cols: list[str]) -> int | None:
    """Find the header row that contains all required column names."""
    for r in range(1, min(10, ws.max_row + 1)):
        found = set()
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(r, c).value or "").strip()
            for req in required_cols:
                if req in h:
                    found.add(req)
        if len(found) >= len(required_cols) * 0.6:  # At least 60% match
            return r
    return None


def get_col_map(ws, header_row: int) -> dict[str, int]:
    """Build column name -> column index map."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(header_row, c).value or "").strip()
        if h:
            col_map[h] = c
    return col_map


def get_val(ws, row: int, col_map: dict, *names: str) -> str:
    """Get cell value by column names (tries each name)."""
    for n in names:
        if n in col_map:
            val = ws.cell(row, col_map[n]).value
            return str(val).strip() if val else ""
    return ""


# ============================================================
# Parse a single position sheet
# ============================================================
def parse_position_sheet(ws, sheet_name: str, year: int) -> list[dict]:
    """Parse one sheet of a position table. sheet_name provides city info."""
    positions = []
    if ws.max_row < 5 or ws.max_column < 8:
        return positions

    # Find header
    required_cols = ["单位名称", "职位名称"]
    header_row = find_header_row(ws, required_cols)
    if header_row is None:
        return positions

    col_map = get_col_map(ws, header_row)

    # Determine city
    # Method 1: From sheet name
    city = city_from_sheet_name(sheet_name)
    # Method 2: From "地区" column in data
    if not city and "地区" in col_map:
        # First data row's 地区 column
        for r in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
            c = get_val(ws, r, col_map, "地区")
            if c and city_from_sheet_name(c):
                city = city_from_sheet_name(c)
                break
    # Method 3: Try column 1 (2025/2026 format)
    if not city:
        c1 = str(ws.cell(header_row + 1, 1).value or "").strip()
        if city_from_sheet_name(c1):
            city = city_from_sheet_name(c1)
    if not city:
        city = "省直"  # fallback

    for r in range(header_row + 1, ws.max_row + 1):
        # For 2023/2024 format: city from sheet or column
        row_city = city
        # If column 1 contains city name (2025/2026 format), use that
        c1 = str(ws.cell(r, 1).value or "").strip()
        c1_city = city_from_sheet_name(c1)
        if c1_city and c1_city != "省直":
            row_city = c1_city

        unit = get_val(ws, r, col_map, "单位名称")
        pn = get_val(ws, r, col_map, "职位名称")

        if not unit or not pn:
            continue
        # Skip header-like rows
        if unit == "单位名称" or pn == "职位名称":
            continue

        es = get_val(ws, r, col_map, "笔试考试科目", "申论考试类别", "考试科目")
        category = infer_category(es)

        enr_str = get_val(ws, r, col_map, "招录人数")
        try: enr = int(float(enr_str))
        except: enr = 1

        major = get_val(ws, r, col_map, "专业要求")
        gender = get_val(ws, r, col_map, "性别要求")
        exp = get_val(ws, r, col_map, "基层工作年限要求")
        age = get_val(ws, r, col_map, "最高年龄要求")
        edu = get_val(ws, r, col_map, "最低学历要求")
        deg = get_val(ws, r, col_map, "学位要求")
        ratio = get_val(ws, r, col_map, "面试与考察比例", "面试比例")

        positions.append({
            "year": year,
            "city": row_city,
            "department": unit,
            "position_name": pn,
            "exam_category": category,
            "exam_subject": es or None,
            "education_requirement": edu or "本科及以上",
            "degree_requirement": deg or "学士",
            "major_requirement": major if major and major != "不限" else None,
            "political_requirement": "不限",
            "gender_requirement": gender if gender and gender != "不限" else "不限",
            "experience_requirement": exp if exp and exp != "不限" else "不限",
            "age_limit": age if age else "35周岁以下",
            "enrollment_count": enr,
            "interview_ratio": ratio if ratio else "1:3",
        })

    return positions


# ============================================================
# Download files from topic
# ============================================================
def fetch_topic_xlsx_urls(year: int, topic_id: int) -> list[str]:
    url = f"https://www.hxw.gov.cn/topic/{topic_id}/index.html"
    req = urllib.request.Request(url, headers={"User-Agent": "HunanExamAssistant/1.0"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        html = resp.read().decode("utf-8", errors="replace")
    links = re.findall(r"""href=["']([^"']*\.xlsx?)["']""", html, re.I)
    return [l if l.startswith("http") else
            f"https://img.hxw.gov.cn{l}" if l.startswith("/") else
            f"https://www.hxw.gov.cn/{l}"
            for l in links]


def download(url: str, dest: Path) -> bool:
    if dest.exists():
        return True
    try:
        urllib.request.urlretrieve(url, str(dest))
        return True
    except Exception as e:
        print(f"    Download failed: {e}")
        return False


# ============================================================
# Parse interview score files
# ============================================================
def parse_interview_scores(filepath: Path) -> dict[tuple[str, str], dict]:
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    scores = defaultdict(list)

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 5: continue

        unit_col = pos_col = score_col = None
        # Find columns from any header row
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                h = str(ws.cell(r, c).value or "")
                if ("单位" in h or "招录" in h) and not unit_col: unit_col = c
                if ("职位" in h) and not pos_col: pos_col = c
                if ("成绩" in h or "笔试" in h) and not score_col: score_col = c

        if not unit_col: unit_col = 5
        if not pos_col: pos_col = 6
        if not score_col: score_col = 11

        for r in range(2, ws.max_row + 1):
            unit = str(ws.cell(r, unit_col).value or "").strip()
            pn = str(ws.cell(r, pos_col).value or "").strip()
            sc = ws.cell(r, score_col).value
            if unit and pn and sc:
                try: scores[(unit, pn)].append(float(sc))
                except: pass

    wb.close()
    return {k: {"count": len(v), "min": round(min(v),2), "max": round(max(v),2),
                "avg": round(sum(v)/len(v),2)} for k, v in scores.items()}


# ============================================================
# Main
# ============================================================
def main():
    print("=" * 60)
    print("湖南省考多年数据导入 V3")
    print("=" * 60)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    for year in [2023, 2024, 2025, 2026]:
        print(f"\n{'='*40}")
        print(f"  {year}年")
        print(f"{'='*40}")

        existing = db.query(PositionHistory).filter(PositionHistory.year == year).count()

        topic_id = TOPIC_IDS[year]
        all_urls = fetch_topic_xlsx_urls(year, topic_id)
        print(f"  Found {len(all_urls)} Excel files on topic page")

        # Download all files
        downloaded = []
        for url in all_urls:
            fname = f"{year}_{url.split('/')[-1]}"
            dest = DATA_DIR / fname
            if download(url, dest):
                downloaded.append((dest, url))

        # Parse all files for positions
        all_positions = []
        for filepath, url in downloaded:
            # Skip .xls (old format files)
            if filepath.suffix.lower() == '.xls':
                continue
            try:
                wb = openpyxl.load_workbook(str(filepath), data_only=True)
                file_has_positions = False
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    positions = parse_position_sheet(ws, sn, year)
                    if positions:
                        print(f"    {filepath.name} / {sn}: {len(positions)} positions")
                        all_positions.extend(positions)
                        file_has_positions = True
                wb.close()
            except Exception as e:
                print(f"    Skipping {filepath.name}: {e}")

        if not all_positions:
            print(f"  No positions found, skipping")
            continue

        # Deduplicate by (unit, position_name)
        seen = set()
        unique_positions = []
        for p in all_positions:
            key = (p["department"], p["position_name"])
            if key not in seen:
                seen.add(key)
                unique_positions.append(p)

        print(f"  Total unique positions: {len(unique_positions)}")

        # Try to find interview score data
        # For 2026, we already have it. For other years, look at downloaded files
        score_stats = {}
        for filepath, url in downloaded:
            if filepath.suffix.lower() == '.xls':
                continue
            try:
                # Quick check if this could be an interview file
                wb = openpyxl.load_workbook(str(filepath), data_only=True)
                is_interview = False
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    for r in range(1, min(3, ws.max_row+1)):
                        for c in range(1, min(15, ws.max_column+1)):
                            if "笔试成绩" in str(ws.cell(r,c).value or ""):
                                is_interview = True
                                break
                wb.close()
                if is_interview:
                    stats = parse_interview_scores(filepath)
                    score_stats.update(stats)
                    print(f"    Found interview scores: {filepath.name} -> {len(stats)} positions")
            except:
                pass

        # If not enough scores, try to find interview announcements
        if len(score_stats) < 10 and year < 2026:
            print(f"  Searching for interview announcements...")
            # Fetch topic page and find content links with interview keywords
            url = f"https://www.hxw.gov.cn/topic/{topic_id}/index.html"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "HunanExamAssistant/1.0"})
                with urllib.request.urlopen(req, timeout=15) as resp:
                    html = resp.read().decode("utf-8", errors="replace")
                content_links = list(set(re.findall(
                    r"""href=["'](/content/\d{4}/\d{2}/\d{2}/\d+\.html)["']""", html
                )))
                print(f"    Found {len(content_links)} content pages")

                for cl in content_links[:30]:
                    page_url = f"https://www.hxw.gov.cn{cl}"
                    try:
                        time.sleep(0.3)
                        req2 = urllib.request.Request(page_url, headers={"User-Agent": "HunanExamAssistant/1.0"})
                        with urllib.request.urlopen(req2, timeout=10) as resp2:
                            page_html = resp2.read().decode("utf-8", errors="replace")
                        title = re.search(r'<title>([^<]+)</title>', page_html)
                        title_text = title.group(1) if title else ""

                        if any(kw in title_text for kw in ["面试", "资格复审", "资格审查", "体能测评", "进面", "入围", "分数线"]):
                            xlsx_links = re.findall(r"""href=["']([^"']*\.xlsx?)["']""", page_html, re.I)
                            for xl in xlsx_links:
                                full = xl if xl.startswith("http") else \
                                       f"https://img.hxw.gov.cn{xl}" if xl.startswith("/") else \
                                       f"https://www.hxw.gov.cn/{xl}"
                                fname = f"{year}_interview_{abs(hash(full))}.xlsx"
                                dest = DATA_DIR / fname
                                if download(full, dest):
                                    try:
                                        stats = parse_interview_scores(dest)
                                        old_count = len(score_stats)
                                        score_stats.update(stats)
                                        new = len(score_stats) - old_count
                                        if new > 0:
                                            print(f"      +{new} scores from {title_text[:40]}")
                                    except:
                                        pass
                    except:
                        pass
            except:
                pass

        # If we have old data, delete and re-import
        if existing > 0:
            db.query(PositionHistory).filter(PositionHistory.year == year).delete()
            print(f"  Deleted {existing} old {year} records")

        # Import
        imported = 0
        with_scores = 0
        for p in unique_positions:
            key = (p["department"], p["position_name"])
            sc = score_stats.get(key, {})
            if not sc:
                for k, v in score_stats.items():
                    if p["department"] in k[0] or k[0] in p["department"]:
                        if p["position_name"] in k[1] or k[1] in p["position_name"]:
                            sc = v; break

            if sc:
                with_scores += 1

            db.add(PositionHistory(
                year=p["year"], city=p["city"],
                department=p["department"], position_name=p["position_name"],
                exam_category=p["exam_category"],
                exam_subject=p.get("exam_subject"),
                education_requirement=p["education_requirement"],
                degree_requirement=p["degree_requirement"],
                major_requirement=p["major_requirement"],
                political_requirement=p.get("political_requirement", "不限"),
                gender_requirement=p["gender_requirement"],
                experience_requirement=p["experience_requirement"],
                age_limit=p["age_limit"],
                enrollment_count=p["enrollment_count"],
                interview_ratio=p["interview_ratio"],
                min_score_interview=sc.get("min"),
                max_score_interview=sc.get("max"),
                avg_score_interview=sc.get("avg"),
                source=f"红星网 topic/{topic_id}",
            ))
            imported += 1

        db.commit()
        print(f"  Imported: {imported} positions ({with_scores} with scores)")

    # Final stats
    from sqlalchemy import func, text
    total = db.query(PositionHistory).count()
    print(f"\n{'='*40}")
    print(f"Total positions in DB: {total}")
    for y in [2023, 2024, 2025, 2026]:
        cnt = db.query(PositionHistory).filter(PositionHistory.year == y).count()
        scored = db.query(PositionHistory).filter(
            PositionHistory.year == y, PositionHistory.min_score_interview.isnot(None)
        ).count()
        cats = db.query(PositionHistory.exam_category, func.count(PositionHistory.id))\
            .filter(PositionHistory.year == y).group_by(PositionHistory.exam_category).all()
        cat_str = ", ".join(f"{c[0]}={c[1]}" for c in cats)
        print(f"  {y}: {cnt} positions ({scored} scored) [{cat_str}]")
    db.close()


if __name__ == "__main__":
    main()
