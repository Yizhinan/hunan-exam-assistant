"""
完整导入 2023-2026 湖南省考职位数据和面试分数。
每个年度：下载所有Excel → 识别职位表 → 解析岗位 → 匹配面试分 → 入库

用法: cd backend && python scripts/import_all_years_v2.py
"""

import os, sys, re, urllib.request, time
from pathlib import Path
from collections import defaultdict

os.chdir(Path(__file__).resolve().parent.parent)
sys.path.insert(0, ".")

from app.core.database import engine, Base, SessionLocal
from app.models.position import PositionHistory
import app.models.user, app.models.document, app.models.essay, app.models.daily_essay  # noqa

import openpyxl

DATA_DIR = Path("D:/tmp/hunan_exam_data")
DATA_DIR.mkdir(exist_ok=True)

TOPIC_IDS = {2023: 227, 2024: 247, 2025: 267, 2026: 294}

# ============================================================
# Step 1: Download all Excel files from each topic page
# ============================================================
def fetch_topic_files(year: int, topic_id: int) -> list[str]:
    """Fetch topic page and return all .xlsx/.xls URLs."""
    url = f"https://www.hxw.gov.cn/topic/{topic_id}/index.html"
    print(f"  Fetching {url}...")
    req = urllib.request.Request(url, headers={"User-Agent": "HunanExamAssistant/1.0"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    links = re.findall(r"""href=["']([^"']*\.xlsx?)["']""", html, re.I)
    result = []
    for link in links:
        full = link if link.startswith("http") else \
               f"https://img.hxw.gov.cn{link}" if link.startswith("/") else \
               f"https://www.hxw.gov.cn/{link}"
        result.append(full)
    return result


def download(url: str, dest: Path) -> bool:
    if dest.exists():
        return True
    try:
        print(f"    Downloading {url.split('/')[-1][:60]}...")
        urllib.request.urlretrieve(url, str(dest))
        return True
    except Exception as e:
        print(f"    Download failed: {e}")
        return False


# ============================================================
# Step 2: Identify position table vs other files
# ============================================================
def is_position_table(filepath: Path) -> bool:
    """Check if an Excel file is a position table (has 单位名称 column)."""
    try:
        if filepath.suffix.lower() == '.xls':
            return False  # Old format, skip (usually 专业目录 etc)
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.max_column < 10 or ws.max_row < 20:
                continue
            # Check first 5 rows for "单位名称" header
            for r in range(1, min(6, ws.max_row + 1)):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(r, c).value or "")
                    if "单位名称" in val:
                        wb.close()
                        return True
        wb.close()
    except:
        pass
    return False


def is_interview_file(filepath: Path) -> bool:
    """Check if file contains interview score data."""
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            for r in range(1, min(3, ws.max_row + 1)):
                for c in range(1, min(15, ws.max_column + 1)):
                    val = str(ws.cell(r, c).value or "")
                    if "笔试成绩" in val or "面试成绩" in val:
                        wb.close()
                        return True
        wb.close()
    except:
        pass
    return False


# ============================================================
# Step 3: Parse position table
# ============================================================
def infer_category(exam_subject: str) -> str:
    if not exam_subject:
        return "省市直"
    if "行政执法" in exam_subject:
        return "行政执法"
    if "县乡" in exam_subject:
        return "县乡基层"
    if "省市" in exam_subject:
        return "省市直"
    return "省市直"


def parse_positions(filepath: Path, year: int) -> list[dict]:
    """Parse a position table Excel and return list of position dicts."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    positions = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_column < 10 or ws.max_row < 20:
            continue

        # Find header row
        header_row = None
        for r in range(1, min(8, ws.max_row + 1)):
            c1 = str(ws.cell(r, 1).value or "").strip()
            c2 = str(ws.cell(r, 2).value or "").strip()
            if c1 in ("地区", "说明") or c2 == "单位名称":
                header_row = r
                break
            # Also try: first row with "单位名称" anywhere
            for c in range(1, min(25, ws.max_column + 1)):
                if "单位名称" in str(ws.cell(r, c).value or ""):
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            print(f"    Skipping sheet '{sn}': no header with 单位名称 found")
            continue

        # Build column map
        col_map = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(header_row, c).value or "").strip()
            if h:
                col_map[h] = c

        def get_col(row, *names):
            for n in names:
                if n in col_map:
                    val = ws.cell(row, col_map[n]).value
                    return str(val).strip() if val else ""
            return ""

        for r in range(header_row + 1, ws.max_row + 1):
            city = str(ws.cell(r, 1).value or "").strip()
            unit = get_col(r, "单位名称")
            pn = get_col(r, "职位名称")

            if not unit or not pn:
                continue
            if city in ("地区", "说明", "") and unit == "单位名称":
                continue

            es = get_col(r, "笔试考试科目")
            category = infer_category(es)

            enr_str = get_col(r, "招录人数")
            try:
                enr = int(float(enr_str))
            except:
                enr = 1

            major = get_col(r, "专业要求")
            gender = get_col(r, "性别要求")
            exp = get_col(r, "基层工作年限要求")
            age = get_col(r, "最高年龄要求")
            edu = get_col(r, "最低学历要求")
            deg = get_col(r, "学位要求")
            ratio = get_col(r, "面试与考察比例")
            identity = get_col(r, "报考人员身份要求")

            positions.append({
                "year": year,
                "city": city,
                "department": unit,
                "position_name": pn,
                "exam_category": category,
                "exam_subject": es,
                "education_requirement": edu or "本科及以上",
                "degree_requirement": deg or "学士",
                "major_requirement": major if major and major != "不限" else None,
                "political_requirement": "不限",
                "gender_requirement": gender if gender and gender != "不限" else "不限",
                "experience_requirement": exp if exp and exp != "不限" else "不限",
                "age_limit": age if age else "35周岁以下",
                "applicant_identity": identity if identity else "不限",
                "enrollment_count": enr,
                "interview_ratio": ratio if ratio else "1:3",
            })

    wb.close()
    return positions


# ============================================================
# Step 4: Parse interview score table
# ============================================================
def parse_interview_scores(filepath: Path) -> dict[tuple[str, str], dict]:
    """Parse interview Excel, return {(unit, position): {min, max, avg, count}}."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    scores = defaultdict(list)

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 5:
            continue

        # Try to detect columns from header row
        unit_col = pos_col = score_col = None
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                h = str(ws.cell(r, c).value or "")
                if "单位" in h and not unit_col:
                    unit_col = c
                if "职位" in h and not pos_col:
                    pos_col = c
                if ("成绩" in h or "笔试" in h) and not score_col:
                    score_col = c
            if unit_col and pos_col and score_col:
                break

        # Fallbacks
        if not unit_col:
            unit_col = 5
        if not pos_col:
            pos_col = 6
        if not score_col:
            score_col = 11

        header_row = 1
        # Find actual header
        for r in range(1, min(5, ws.max_row + 1)):
            if str(ws.cell(r, unit_col).value or "") not in ("", "None"):
                header_row = r
                break

        for r in range(header_row + 1, ws.max_row + 1):
            unit = str(ws.cell(r, unit_col).value or "").strip()
            pn = str(ws.cell(r, pos_col).value or "").strip()
            sc = ws.cell(r, score_col).value
            if unit and pn and sc:
                try:
                    scores[(unit, pn)].append(float(sc))
                except:
                    pass

    wb.close()
    return {
        k: {
            "count": len(v),
            "min": round(min(v), 2),
            "max": round(max(v), 2),
            "avg": round(sum(v) / len(v), 2),
        }
        for k, v in scores.items()
    }


# ============================================================
# Step 5: Find interview-related content pages
# ============================================================
def find_interview_excels(topic_id: int, year: int) -> list[str]:
    """Crawl content pages to find interview score announcements."""
    url = f"https://www.hxw.gov.cn/topic/{topic_id}/index.html"
    req = urllib.request.Request(url, headers={"User-Agent": "HunanExamAssistant/1.0"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    # Find all content links
    content_links = list(set(re.findall(
        r"""href=["'](/content/\d{4}/\d{2}/\d{2}/\d+\.html)["']""", html
    )))

    # Also look for direct xlsx links that might be interview files
    all_xlsx = re.findall(r"""href=["']([^"']*\.xlsx?)["']""", html, re.I)
    interview_urls = []

    # First check content pages for interview announcements
    for cl in content_links[:30]:
        page_url = f"https://www.hxw.gov.cn{cl}"
        try:
            time.sleep(0.3)
            req2 = urllib.request.Request(page_url, headers={"User-Agent": "HunanExamAssistant/1.0"})
            with urllib.request.urlopen(req2, timeout=10) as resp2:
                page_html = resp2.read().decode("utf-8", errors="replace")

            title = re.search(r'<title>([^<]+)</title>', page_html)
            title_text = title.group(1) if title else ""

            is_interview = any(kw in title_text for kw in
                ["面试", "资格复审", "资格审查", "体能测评", "分数线", "进面", "入围"])

            if is_interview:
                xlsx_links = re.findall(r"""href=["']([^"']*\.xlsx?)["']""", page_html, re.I)
                for xl in xlsx_links:
                    full = xl if xl.startswith("http") else \
                           f"https://img.hxw.gov.cn{xl}" if xl.startswith("/") else \
                           f"https://www.hxw.gov.cn/{xl}"
                    interview_urls.append(full)
                    print(f"    Interview file: {title_text[:50]} -> {full.split('/')[-1][:50]}")
        except:
            pass

    return interview_urls


# ============================================================
# Main
# ============================================================
def main():
    print("=" * 60)
    print("湖南省考多年数据导入 V2")
    print("=" * 60)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    for year in [2023, 2024, 2025, 2026]:
        print(f"\n{'='*40}")
        print(f"  {year}年")
        print(f"{'='*40}")

        # Skip if already imported
        existing = db.query(PositionHistory).filter(PositionHistory.year == year).count()
        if existing > 50:
            print(f"  Already have {existing} positions, skipping")
            continue

        topic_id = TOPIC_IDS[year]
        try:
            # Download all Excel files
            all_urls = fetch_topic_files(year, topic_id)
            print(f"  Found {len(all_urls)} Excel files")

            downloaded = []
            for url in all_urls:
                fname = f"{year}_{url.split('/')[-1]}"
                dest = DATA_DIR / fname
                if download(url, dest):
                    downloaded.append(dest)

            # Identify position table
            pos_file = None
            interview_files = []
            for f in downloaded:
                if is_position_table(f):
                    pos_file = f
                    print(f"  Position table: {f.name}")
                elif is_interview_file(f):
                    interview_files.append(f)
                    print(f"  Interview file: {f.name}")

            if not pos_file:
                print(f"  No position table found among {len(downloaded)} files, skipping")
                continue

            # Parse positions
            positions = parse_positions(pos_file, year)
            print(f"  Parsed {len(positions)} positions")

            # Find interview score data
            interview_urls = find_interview_excels(topic_id, year)
            score_stats = {}

            # Try attached interview files first
            for f in interview_files:
                try:
                    stats = parse_interview_scores(f)
                    score_stats.update(stats)
                    print(f"    +{len(stats)} scores from {f.name}")
                except Exception as e:
                    print(f"    Parse error: {e}")

            # Then try downloaded interview excels
            for iurl in interview_urls[:10]:
                fname = f"{year}_interview_{abs(hash(iurl))}.xlsx"
                dest = DATA_DIR / fname
                if download(iurl, dest):
                    try:
                        stats = parse_interview_scores(dest)
                        score_stats.update(stats)
                        print(f"    +{len(stats)} scores from interview file")
                    except Exception as e:
                        print(f"    Parse error: {e}")

            # Merge and import
            imported = 0
            with_scores = 0
            for p in positions:
                key = (p["department"], p["position_name"])
                sc = score_stats.get(key, {})

                # Fuzzy match
                if not sc:
                    for k, v in score_stats.items():
                        if p["department"] in k[0] or k[0] in p["department"]:
                            if p["position_name"] in k[1] or k[1] in p["position_name"]:
                                sc = v
                                break

                if sc:
                    with_scores += 1

                db.add(PositionHistory(
                    year=p["year"], city=p["city"],
                    department=p["department"], position_name=p["position_name"],
                    exam_category=p["exam_category"],
                    exam_subject=p.get("exam_subject"),
                    education_requirement=p["education_requirement"],
                    degree_requirement=p["degree_requirement"],
                    major_requirement=p["major_requirement"],
                    political_requirement=p.get("political_requirement", "不限"),
                    gender_requirement=p["gender_requirement"],
                    experience_requirement=p["experience_requirement"],
                    age_limit=p["age_limit"],
                    enrollment_count=p["enrollment_count"],
                    interview_ratio=p["interview_ratio"],
                    min_score_interview=sc.get("min"),
                    max_score_interview=sc.get("max"),
                    avg_score_interview=sc.get("avg"),
                    source=f"红星网 topic/{topic_id}",
                ))
                imported += 1

            db.commit()
            print(f"  Imported: {imported} positions ({with_scores} with scores)")

        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()

    # Final stats
    total = db.query(PositionHistory).count()
    print(f"\n{'='*40}")
    print(f"Total positions in DB: {total}")
    years = db.execute(
        "SELECT year, count(*) FROM position_history GROUP BY year ORDER BY year"
    ).fetchall()
    for y, c in years:
        scored = db.query(PositionHistory).filter(
            PositionHistory.year == y, PositionHistory.min_score_interview.isnot(None)
        ).count()
        print(f"  {y}: {c} positions ({scored} with scores)")
    db.close()


if __name__ == "__main__":
    main()
